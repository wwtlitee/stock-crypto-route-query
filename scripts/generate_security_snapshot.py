import csv
import html
import io
import json
import re
import sys
import time
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path
from tempfile import gettempdir

from openpyxl import load_workbook


SCRIPT_PATH = Path(__file__).resolve()
ROOT = SCRIPT_PATH.parents[1] if SCRIPT_PATH.parent.name == "scripts" else SCRIPT_PATH.parents[2]
OUTPUT = ROOT / "assets" / "market-data.js"

NASDAQ_LISTED_URL = "https://www.nasdaqtrader.com/dynamic/SymDir/nasdaqlisted.txt"
OTHER_LISTED_URL = "https://www.nasdaqtrader.com/dynamic/SymDir/otherlisted.txt"
HKEX_LIST_URL = "https://www.hkex.com.hk/eng/services/trading/securities/securitieslists/ListOfSecurities.xlsx"
HKEX_EQUITIES_PAGE_URL = "https://www.hkex.com.hk/Market-Data/Securities-Prices/Equities?sc_lang=en"
HKEX_EQUITY_FILTER_URL = "https://www1.hkex.com.hk/hkexwidget/data/getequityfilter"
EURONEXT_EQUITIES_PAGE_URL = "https://live.euronext.com/en/products/equities/list"
EURONEXT_STOCKS_FALLBACK_URL = (
    "https://live.euronext.com/product_directory/data/stocks-all-places/download?"
    "mics=ALXB%2CALXL%2CALXP%2CENXB%2CENXL%2CMERK%2CMLXB%2CTNLA%2CTNLB%2C"
    "XAMC%2CXAMS%2CXATL%2CXBRU%2CXESM%2CXLDN%2CXLIS%2CXMLI%2CXMSM%2CXOAS%2C"
    "XOSL%2CXPAR%2CXPMC"
)
XETRA_INSTRUMENTS_PAGE_URL = "https://www.cashmarket.deutsche-boerse.com/cash-en/trading/Tradable-Instruments-Xetra/"
XETRA_ALL_TRADABLE_FALLBACK_URL = (
    "https://www.cashmarket.deutsche-boerse.com/resource/blob/1528/"
    "249d444f2e45fd4c7f8398e87dd03faa/data/t7-xetr-allTradableInstruments.csv"
)
SIX_EQUITY_ISSUERS_URL = "https://www.six-group.com/sheldon/equity_issuers/v1/equity_issuers.json"
BME_LISTED_COMPANIES_PAGE_URL = "https://www.bolsasymercados.es/en/bme-exchange/prices-and-markets/shares/listed-companies.html"
BME_MARKET_ENDPOINT = "https://apiweb.bolsasymercados.es/Market/"
LSE_AIM_INSTRUMENTS_URL = "https://api.londonstockexchange.com/api/gw/lse/download/directories/aim-adviser/instruments"
NASDAQ_NORDIC_SCREENER_URL = "https://api.nasdaq.com/api/nordic/screener/shares?category={category}&tableonly=false&lang=en"
BINANCE_RWA_URL = (
    "https://www.binance.com/bapi/defi/v1/public/wallet-direct/buw/wallet/market/"
    "token/rwa/stock/detail/list/ai"
)
YAHOO_CHART_URL = "https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?range=5d&interval=1d"
YAHOO_SEARCH_URL = "https://query1.finance.yahoo.com/v1/finance/search?q={symbol}&quotesCount=4&newsCount=5&lang=en-US&region=US"


BROKER_PLATFORMS = [
    {
        "id": "ibkr",
        "name": "Interactive Brokers",
        "short": "IB",
        "logo": "./assets/logos/ibkr.png",
        "url": "https://www.interactivebrokers.com/en/trading/products-stocks.php",
        "supportedMarkets": ["US", "HK", "STOCKHOLM", "NASDAQ_NORDIC", "EURONEXT", "XETRA", "SIX", "BME", "LSE_AIM", "OTC"],
        "requiresLoginMarkets": [],
        "source": "IBKR public products and contract search materials",
    },
    {
        "id": "saxobank",
        "name": "Saxo Bank",
        "short": "SX",
        "logo": "./assets/logos/saxobank.png",
        "url": "https://www.home.saxo/products/stocks",
        "supportedMarkets": ["US", "HK", "STOCKHOLM", "NASDAQ_NORDIC", "EURONEXT", "XETRA", "SIX", "BME", "LSE_AIM"],
        "requiresLoginMarkets": ["OTC"],
        "source": "Saxo OpenAPI reference data and product materials",
    },
    {
        "id": "futu",
        "name": "富途牛牛",
        "short": "FT",
        "logo": "./assets/logos/futu.png",
        "url": "https://www.futuhk.com/",
        "supportedMarkets": ["US", "HK"],
        "requiresLoginMarkets": [],
        "source": "Futu public market scope",
    },
    {
        "id": "moomoo",
        "name": "moomoo",
        "short": "MM",
        "logo": "./assets/logos/moomoo.png",
        "url": "https://www.moomoo.com/",
        "supportedMarkets": ["US", "HK"],
        "requiresLoginMarkets": [],
        "source": "moomoo public market scope",
    },
    {
        "id": "tiger",
        "name": "老虎证券",
        "short": "TG",
        "logo": "./assets/logos/tiger.png",
        "url": "https://www.itigerup.com/",
        "supportedMarkets": ["US", "HK"],
        "requiresLoginMarkets": [],
        "source": "Tiger Brokers public market scope",
    },
    {
        "id": "longbridge",
        "name": "Longbridge",
        "short": "LB",
        "logo": "./assets/logos/longbridge.png",
        "url": "https://longbridge.com/",
        "supportedMarkets": ["US", "HK"],
        "requiresLoginMarkets": [],
        "source": "Longbridge public market scope",
    },
    {
        "id": "webull",
        "name": "Webull",
        "short": "WB",
        "logo": "./assets/logos/webull.png",
        "url": "https://www.webull.com/",
        "supportedMarkets": ["US"],
        "requiresLoginMarkets": ["OTC"],
        "source": "Webull public market scope",
    },
    {
        "id": "schwab",
        "name": "Charles Schwab",
        "short": "CS",
        "logo": "./assets/logos/schwab.png",
        "url": "https://www.schwab.com/",
        "supportedMarkets": ["US", "OTC"],
        "requiresLoginMarkets": [],
        "source": "Schwab public market scope",
    },
    {
        "id": "robinhood",
        "name": "Robinhood",
        "short": "RH",
        "logo": "./assets/logos/robinhood.png",
        "url": "https://robinhood.com/us/en/",
        "supportedMarkets": ["US"],
        "requiresLoginMarkets": [],
        "source": "Robinhood public market scope",
    },
    {
        "id": "firstrade",
        "name": "Firstrade",
        "short": "FD",
        "logo": "./assets/logos/firstrade.png",
        "url": "https://www.firstrade.com/",
        "supportedMarkets": ["US", "OTC"],
        "requiresLoginMarkets": [],
        "source": "Firstrade public market scope",
    },
    {
        "id": "trading212",
        "name": "Trading 212",
        "short": "T2",
        "logo": "./assets/logos/trading212.png",
        "url": "https://www.trading212.com/",
        "supportedMarkets": ["US", "EURONEXT", "XETRA"],
        "requiresLoginMarkets": ["STOCKHOLM", "OTC"],
        "source": "Trading 212 public product scope",
    },
    {
        "id": "etoro",
        "name": "eToro",
        "short": "ET",
        "logo": "./assets/logos/etoro.png",
        "url": "https://www.etoro.com/",
        "supportedMarkets": ["US", "EURONEXT", "XETRA"],
        "requiresLoginMarkets": ["STOCKHOLM", "OTC"],
        "source": "eToro public product scope",
    },
    {
        "id": "scalable-capital",
        "name": "Scalable Capital",
        "short": "SC",
        "logo": "./assets/logos/scalable_capital.png",
        "url": "https://de.scalable.capital/en/trading",
        "supportedMarkets": [],
        "requiresLoginMarkets": [],
        "routeCandidateMarkets": ["EURONEXT", "XETRA", "SIX", "BME", "LSE_AIM", "NASDAQ_NORDIC", "STOCKHOLM"],
        "routeCandidateVenues": ["gettex", "Xetra"],
        "routeCandidateSource": "Scalable Capital trading venue/product materials",
        "source": "Scalable Capital public trading venue/product materials",
    },
    {
        "id": "trade-republic",
        "name": "Trade Republic",
        "short": "TR",
        "logo": "./assets/logos/trade_republic.svg",
        "url": "https://traderepublic.com/",
        "supportedMarkets": [],
        "requiresLoginMarkets": [],
        "routeCandidateMarkets": ["EURONEXT", "XETRA", "SIX", "BME", "LSE_AIM", "NASDAQ_NORDIC", "STOCKHOLM"],
        "routeCandidateVenues": ["Lang & Schwarz Exchange"],
        "routeCandidateSource": "Trade Republic public trading venue/product materials",
        "source": "Trade Republic public trading venue/product materials",
    },
]


MAJOR_STOCK_PERP_SYMBOLS = ["AAPL", "AMZN", "COIN", "GOOGL", "META", "MSFT", "MSTR", "NVDA", "TSLA"]


CRYPTO_PLATFORMS = [
    {
        "id": "binance-rwa",
        "name": "Binance Web3 / RWA",
        "short": "BN",
        "logo": "./assets/logos/binance.png",
        "type": "cex",
        "url": "https://www.binance.com/en/web3wallet",
        "coverageKinds": ["ondo", "xstocks"],
        "source": "Binance Web3 RWA public API",
    },
    {
        "id": "okx-ondo",
        "name": "OKX xAssets / Ondo",
        "short": "OK",
        "logo": "./assets/logos/okx.png",
        "type": "cex",
        "url": "https://www.okx.com/learn/ondo-tokenized-stocks",
        "coverageKinds": ["ondo"],
        "source": "OKX xAssets/Ondo tokenized stocks product page",
    },
    {
        "id": "bitget-stock",
        "name": "Bitget Stock Tokens / Perps",
        "short": "BG",
        "logo": "./assets/logos/bitget.png",
        "type": "cex",
        "url": "https://www.bitget.com/academy/how-to-trade-nvidia-stock-using-usdt-on-bitget-2026-guide",
        "coverageKinds": ["stock_perp"],
        "majorSymbols": MAJOR_STOCK_PERP_SYMBOLS,
        "symbolTemplate": "{ticker}on / {ticker}USDT",
        "marketLabel": "代币化/合约敞口",
        "routeLabel": "代币化股票/股票永续",
        "productLabel": "代币化股票/股票合约",
        "source": "Bitget tokenized stock and stock perpetuals product materials",
    },
    {
        "id": "bybit-xstocks",
        "name": "Bybit xStocks",
        "short": "BY",
        "logo": "./assets/logos/bybit.png",
        "type": "cex",
        "url": "https://www.bybit.com/en/help-center/article/FAQ-xStocks-on-Bybit",
        "coverageKinds": ["xstocks"],
        "source": "Bybit xStocks FAQ",
    },
    {
        "id": "kraken-xstocks",
        "name": "Kraken xStocks",
        "short": "KR",
        "logo": "./assets/logos/kraken.png",
        "type": "cex",
        "url": "https://www.kraken.com/xstocks",
        "coverageKinds": ["xstocks"],
        "source": "Kraken xStocks FAQ",
    },
    {
        "id": "gate-xstocks",
        "name": "Gate xStocks",
        "short": "GT",
        "logo": "./assets/logos/gate.png",
        "type": "cex",
        "url": "https://www.gate.com/",
        "coverageKinds": ["xstocks"],
        "source": "Gate xStocks product materials",
    },
    {
        "id": "trade-xyz",
        "name": "trade.xyz",
        "short": "TX",
        "logo": "./assets/logos/trade_xyz.png",
        "type": "dex",
        "url": "https://trade.xyz/",
        "coverageKinds": ["xstocks"],
        "source": "xStocks ecosystem trading entry",
    },
    {
        "id": "hyperliquid",
        "name": "Hyperliquid HIP-3",
        "short": "HL",
        "logo": "./assets/logos/hyperliquid.png",
        "type": "dex",
        "url": "https://app.hyperliquid.xyz/trade",
        "coverageKinds": ["stock_perp"],
        "majorSymbols": MAJOR_STOCK_PERP_SYMBOLS,
        "source": "Hyperliquid stock perpetual discovery entry",
    },
    {
        "id": "aster",
        "name": "Aster",
        "short": "AS",
        "logo": "./assets/logos/aster.png",
        "type": "dex",
        "url": "https://www.asterdex.com/",
        "coverageKinds": ["stock_perp"],
        "majorSymbols": MAJOR_STOCK_PERP_SYMBOLS,
        "source": "Aster stock perpetual discovery entry",
    },
]


PROFILE_OVERRIDES = {
    "AAPL": {"name": "Apple", "aliases": ["苹果"], "description": "Apple Inc. 是消费电子、软件与服务公司，核心产品包括 iPhone、Mac、iPad、Apple Watch 与服务订阅。"},
    "TSLA": {"name": "Tesla", "aliases": ["特斯拉"], "description": "Tesla, Inc. 是电动汽车、能源存储与自动驾驶技术公司，主要业务覆盖汽车销售、能源产品和软件服务。"},
    "NVDA": {"name": "NVIDIA", "aliases": ["英伟达", "英伟達"], "description": "NVIDIA Corporation 是 GPU、AI 加速计算和数据中心芯片公司，也是 AI 基础设施的重要供应商。"},
    "BABA": {"name": "Alibaba", "aliases": ["阿里", "阿里巴巴", "9988", "9988.HK"], "relatedListings": ["9988.HK"]},
    "9988.HK": {"name": "Alibaba Group", "aliases": ["阿里", "阿里巴巴", "BABA"], "relatedListings": ["BABA"]},
    "SIVE.ST": {
        "name": "Sivers Semiconductors AB",
        "aliases": ["SIVE", "STO:SIVE", "SIVEF", "Sivers"],
        "description": "Sivers Semiconductors AB 是瑞典半导体公司，主上市地为 Nasdaq Stockholm。",
        "relatedListings": ["SIVEF"],
        "researchSymbol": "SIVE.ST",
        "researchKeywords": ["sivers", "sive", "semiconductors"],
        "verifiedBrokerRoutes": [
            {
                "platformId": "saxobank",
                "listingSymbol": "SIVE:xome",
                "status": "verified_tradable",
                "sourceType": "official_instrument_page",
                "sourceUrl": "https://www.home.saxo/markets/stocks/sive-xome",
                "evidence": "Saxo 官方 Sivers Semiconductors AB 标的页 FAQ 写明可在 SaxoInvestor 与 SaxoTrader 交易，并列出 ticker SIVE:xome。",
                "verifiedAt": "2026-06-04",
            },
            {
                "platformId": "etoro",
                "listingSymbol": "SIVE.ST",
                "status": "verified_tradable",
                "sourceType": "official_instrument_page",
                "sourceUrl": "https://www.etoro.com/markets/sive.st",
                "evidence": "eToro 官方 SIVE.ST 标的页显示 Sivers Semiconductors AB 并引导用户购买该股票。",
                "verifiedAt": "2026-06-04",
            },
            {
                "platformId": "webull",
                "listingSymbol": "OTCPK:SIVEF",
                "status": "verified_tradable",
                "sourceType": "official_instrument_page",
                "sourceUrl": "https://www.webull.com/quote/otcpk-sivef",
                "evidence": "Webull 官方 OTCPK:SIVEF 标的页出现 Trade SIVEF，并展示 Sivers Semiconductors AB 股票信息。",
                "verifiedAt": "2026-06-04",
            },
        ],
        "newsProbeNote": "free-api-discovery 对 SIVE 的泛财经新闻命中噪音高，本版本仅展示强相关资讯诊断，不把泛商业新闻推到前台。",
    },
    "SIVEF": {
        "name": "Sivers Semiconductors AB",
        "aliases": ["SIVE", "SIVE.ST", "STO:SIVE", "Sivers"],
        "description": "SIVEF 是 Sivers Semiconductors AB 的美国 OTC 关联代码，主上市代码为 SIVE.ST。",
        "relatedListings": ["SIVE.ST"],
        "researchSymbol": "SIVE.ST",
        "researchKeywords": ["sivers", "sive", "semiconductors"],
        "verifiedBrokerRoutes": [
            {
                "platformId": "webull",
                "listingSymbol": "OTCPK:SIVEF",
                "status": "verified_tradable",
                "sourceType": "official_instrument_page",
                "sourceUrl": "https://www.webull.com/quote/otcpk-sivef",
                "evidence": "Webull 官方 OTCPK:SIVEF 标的页出现 Trade SIVEF，并展示 Sivers Semiconductors AB 股票信息。",
                "verifiedAt": "2026-06-04",
            },
            {
                "platformId": "saxobank",
                "listingSymbol": "SIVE:xome",
                "status": "verified_tradable",
                "sourceType": "official_instrument_page",
                "sourceUrl": "https://www.home.saxo/markets/stocks/sive-xome",
                "evidence": "Saxo 官方 Sivers Semiconductors AB 标的页 FAQ 写明可在 SaxoInvestor 与 SaxoTrader 交易，并列出 ticker SIVE:xome。",
                "verifiedAt": "2026-06-04",
            },
            {
                "platformId": "etoro",
                "listingSymbol": "SIVE.ST",
                "status": "verified_tradable",
                "sourceType": "official_instrument_page",
                "sourceUrl": "https://www.etoro.com/markets/sive.st",
                "evidence": "eToro 官方 SIVE.ST 标的页显示 Sivers Semiconductors AB 并引导用户购买该股票。",
                "verifiedAt": "2026-06-04",
            },
        ],
        "newsProbeNote": "free-api-discovery 对 SIVE 的泛财经新闻命中噪音高，本版本仅展示强相关资讯诊断，不把泛商业新闻推到前台。",
    },
}


def download_text(url):
    request = urllib.request.Request(url, headers={"User-Agent": "TradeRouteSnapshot/2.0"})
    with urllib.request.urlopen(request, timeout=60) as response:
        return response.read().decode("utf-8", errors="replace")


def download_json(url, user_agent="TradeRouteSnapshot/2.1"):
    request = urllib.request.Request(url, headers={"User-Agent": user_agent})
    with urllib.request.urlopen(request, timeout=60) as response:
        return json.loads(response.read().decode("utf-8", errors="replace"))


def download_json_with_headers(url, headers=None, timeout=60, retries=1, retry_delay=0.5):
    last_error = None
    for attempt in range(retries):
        try:
            request = urllib.request.Request(url, headers=headers or {"User-Agent": "TradeRouteSnapshot/2.4"})
            with urllib.request.urlopen(request, timeout=timeout) as response:
                return json.loads(response.read().decode("utf-8", errors="replace"))
        except Exception as exc:
            last_error = exc
            if attempt < retries - 1:
                time.sleep(retry_delay * (attempt + 1))
    raise last_error


def download_binary(url, target):
    request = urllib.request.Request(url, headers={"User-Agent": "TradeRouteSnapshot/2.0"})
    with urllib.request.urlopen(request, timeout=90) as response:
        target.write_bytes(response.read())


def clean_name(value):
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text


def call_with_retries(callback, attempts=3, retry_delay=1.0):
    last_error = None
    for attempt in range(attempts):
        try:
            return callback()
        except Exception as exc:
            last_error = exc
            if attempt < attempts - 1:
                time.sleep(retry_delay * (attempt + 1))
    raise last_error


def classify_us_instrument(name, etf_flag):
    lower = name.lower()
    if etf_flag == "Y":
        return "ETF/基金类标的"
    if "warrant" in lower:
        return "权证"
    if "right" in lower:
        return "权利"
    if "unit" in lower:
        return "单位"
    if "preferred" in lower or "depositary shares" in lower:
        return "优先股/存托"
    if "american depositary" in lower or " adr" in lower:
        return "ADR"
    return "普通股"


def normalize_symbol(symbol):
    return str(symbol or "").strip().upper()


def row_to_stock(listing, source_name):
    symbol = listing["symbol"]
    override = PROFILE_OVERRIDES.get(symbol, {})
    name = override.get("name") or listing["name"]
    aliases = set([symbol.lower(), listing["symbolRaw"].lower(), name.lower()])
    for alias in listing.get("aliases", []):
        aliases.add(alias.lower())
    for alias in override.get("aliases", []):
        aliases.add(alias.lower())
    symbols = [symbol]
    for related in override.get("relatedListings", []):
        symbols.append(related)
    return {
        "id": re.sub(r"[^a-z0-9]+", "-", symbol.lower()).strip("-"),
        "name": name,
        "symbols": sorted(set(symbols), key=symbols.index),
        "aliases": sorted(aliases),
        "summary": override.get("description") or f"{name} 是 {listing['exchangeName']} 收录的 {listing['instrumentType']}。",
        "updated": datetime.now().strftime("%Y-%m-%d"),
        "profile": {
            "code": symbol,
            "company": name,
            "description": override.get("description") or f"{name}，主上市市场为 {listing['exchangeName']}。",
            "assetType": listing["instrumentType"],
            "primaryMarket": listing["market"],
            "primaryExchange": listing["exchangeName"],
            "country": listing.get("country", ""),
            "currency": listing.get("currency", ""),
            "relatedListings": override.get("relatedListings", []),
            "source": source_name,
            "verifiedBrokerRoutes": override.get("verifiedBrokerRoutes", []),
        },
        "listings": [listing],
    }


def parse_nasdaq_listed(text):
    rows = []
    for line in text.splitlines()[1:]:
        if not line or line.startswith("File Creation Time"):
            continue
        parts = line.split("|")
        if len(parts) < 8 or parts[3] == "Y":
            continue
        symbol = normalize_symbol(parts[0])
        name = clean_name(parts[1])
        rows.append({
            "symbol": symbol,
            "symbolRaw": symbol,
            "name": name,
            "market": "US",
            "exchange": "NASDAQ",
            "exchangeName": "Nasdaq",
            "country": "US",
            "currency": "USD",
            "instrumentType": classify_us_instrument(name, parts[6]),
            "source": "NasdaqTrader nasdaqlisted",
        })
    return rows


def parse_other_listed(text):
    exchange_map = {"N": ("NYSE", "New York Stock Exchange"), "A": ("NYSE American", "NYSE American"), "P": ("NYSE Arca", "NYSE Arca"), "Z": ("Cboe BZX", "Cboe BZX"), "V": ("IEX", "IEX")}
    rows = []
    for line in text.splitlines()[1:]:
        if not line or line.startswith("File Creation Time"):
            continue
        parts = line.split("|")
        if len(parts) < 8 or parts[6] == "Y":
            continue
        symbol = normalize_symbol(parts[0])
        name = clean_name(parts[1])
        exchange, exchange_name = exchange_map.get(parts[2], (parts[2], parts[2]))
        rows.append({
            "symbol": symbol,
            "symbolRaw": symbol,
            "name": name,
            "market": "US",
            "exchange": exchange,
            "exchangeName": exchange_name,
            "country": "US",
            "currency": "USD",
            "instrumentType": classify_us_instrument(name, parts[4]),
            "source": "NasdaqTrader otherlisted",
        })
    return rows


def parse_hkex(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["ListOfSecurities"]
    rows = []
    for index, row in enumerate(ws.iter_rows(values_only=True), 1):
        if index <= 3:
            continue
        code, name, category, sub_category, _, isin = row[:6]
        if not code or not name or not category:
            continue
        category_text = clean_name(category)
        if category_text not in {"Equity", "Exchange Traded Products", "Real Estate Investment Trusts"}:
            continue
        symbol_raw = str(code).zfill(5)
        symbol = f"{symbol_raw}.HK"
        instrument = "港股"
        if category_text == "Exchange Traded Products":
            instrument = "ETF/基金类标的"
        elif category_text == "Real Estate Investment Trusts":
            instrument = "REIT"
        rows.append({
            "symbol": symbol,
            "symbolRaw": symbol_raw,
            "name": clean_name(name),
            "market": "HK",
            "exchange": "HKEX",
            "exchangeName": "Hong Kong Exchanges and Clearing",
            "country": "HK",
            "currency": "HKD",
            "instrumentType": instrument,
            "isin": clean_name(isin),
            "source": "HKEX List of Securities",
        })
    return rows


def parse_jsonp(raw):
    matched = re.match(r"^[^(]+\((.*)\);?$", raw, re.S)
    if not matched:
        return json.loads(raw)
    return json.loads(matched.group(1))


def fetch_hkex_token():
    html = download_text(HKEX_EQUITIES_PAGE_URL)
    candidates = re.findall(r"return\s+\"([A-Za-z0-9%+/=]+)\"\s*;", html)
    token = next((item for item in candidates if item != "Base64-AES-Encrypted-Token" and len(item) > 40), "")
    if not token:
        raise RuntimeError("HKEX token not found on equities page")
    return urllib.parse.unquote(token)


def parse_hkex_equity_filter(stocklist):
    subtype_map = {
        "ODSH": "港股",
        "DPRC": "存托凭证",
        "PFSH": "优先股/存托",
        "RGHT": "权利",
        "UNIT": "单位",
        "WRNT": "权证",
    }
    rows = []
    for item in stocklist:
        symbol = normalize_symbol(item.get("ric"))
        if not symbol.endswith(".HK"):
            raw_code = str(item.get("sym") or "").strip()
            if not raw_code:
                continue
            symbol = f"{raw_code.zfill(4)}.HK"
        symbol_raw = symbol.replace(".HK", "")
        short_code = str(item.get("sym") or "").strip()
        aliases = [short_code] if short_code and short_code != symbol_raw else []
        rows.append({
            "symbol": symbol,
            "symbolRaw": symbol_raw,
            "aliases": aliases,
            "name": clean_name(item.get("nm")),
            "market": "HK",
            "exchange": "HKEX",
            "exchangeName": "Hong Kong Exchanges and Clearing",
            "country": "HK",
            "currency": clean_name(item.get("ccy")) or "HKD",
            "instrumentType": subtype_map.get(item.get("asset_subtype"), "港交所证券"),
            "source": "HKEX equityfilter all=1",
        })
    return rows


def load_hkex_equity_filter():
    token = fetch_hkex_token()
    qid = int(time.time() * 1000)
    callback = f"jQuery35100000000000000000_{qid}"
    params = {
        "lang": "eng",
        "token": token,
        "sort": "0",
        "order": "1",
        "all": "1",
        "qid": str(qid),
        "callback": callback,
        "_": str(qid),
    }
    url = HKEX_EQUITY_FILTER_URL + "?" + urllib.parse.urlencode(params)
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Referer": HKEX_EQUITIES_PAGE_URL,
        },
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        payload = parse_jsonp(response.read().decode("utf-8", errors="replace"))
    data = payload.get("data") or {}
    if data.get("responsecode") != "000":
        raise RuntimeError(f"HKEX equityfilter failed: {data.get('responsemsg')}")
    return parse_hkex_equity_filter(data.get("stocklist") or [])


EURONEXT_MARKET_SUFFIXES = {
    "Euronext Amsterdam": "AS",
    "Euronext Brussels": "BR",
    "Euronext Dublin": "IR",
    "Euronext Lisbon": "LS",
    "Euronext Paris": "PA",
    "Euronext Growth Amsterdam": "AS",
    "Euronext Growth Brussels": "BR",
    "Euronext Growth Dublin": "IR",
    "Euronext Growth Lisbon": "LS",
    "Euronext Growth Paris": "PA",
    "Euronext Access Brussels": "BR",
    "Euronext Access Lisbon": "LS",
    "Euronext Access Paris": "PA",
    "Euronext Expand Oslo": "OL",
    "Oslo B\u00f8rs": "OL",
    "Euronext Growth Oslo": "OL",
    "Euronext Paris - Multi-currency Trading": "PA",
    "Global Equity Market": "GEM",
    "EuroTLX": "ETLX",
    "Trading After Hours": "TAH",
}


EURONEXT_MARKET_MICS = {
    "Euronext Amsterdam": "XAMS",
    "Euronext Brussels": "XBRU",
    "Euronext Dublin": "XDUB",
    "Euronext Lisbon": "XLIS",
    "Euronext Paris": "XPAR",
    "Euronext Growth Amsterdam": "ALXA",
    "Euronext Growth Brussels": "ALXB",
    "Euronext Growth Dublin": "XDUB",
    "Euronext Growth Lisbon": "ALXL",
    "Euronext Growth Paris": "ALXP",
    "Euronext Access Brussels": "MLXB",
    "Euronext Access Lisbon": "ENXL",
    "Euronext Access Paris": "XPAR",
    "Euronext Expand Oslo": "XOAS",
    "Oslo B\u00f8rs": "XOSL",
    "Euronext Growth Oslo": "MERK",
    "Euronext Paris - Multi-currency Trading": "XPAR",
    "Global Equity Market": "XGEM",
    "EuroTLX": "ETLX",
    "Trading After Hours": "TNLA",
}


def euronext_download_url():
    try:
        html_text = download_text(EURONEXT_EQUITIES_PAGE_URL)
    except Exception:
        return EURONEXT_STOCKS_FALLBACK_URL
    matched = re.search(
        r"(/product_directory\\/data\\/stocks-all-places\\/download\?mics=[^\"'<\s]+)",
        html_text,
    )
    if not matched:
        return EURONEXT_STOCKS_FALLBACK_URL
    raw_url = html.unescape(matched.group(1).replace("\\/", "/"))
    return urllib.parse.urljoin("https://live.euronext.com", raw_url)


def xetra_download_url():
    try:
        html_text = download_text(XETRA_INSTRUMENTS_PAGE_URL)
    except Exception:
        return XETRA_ALL_TRADABLE_FALLBACK_URL
    matched = re.search(r'href="([^"]*t7-xetr-allTradableInstruments\.csv[^"]*)"', html_text)
    if not matched:
        return XETRA_ALL_TRADABLE_FALLBACK_URL
    return urllib.parse.urljoin(XETRA_INSTRUMENTS_PAGE_URL, html.unescape(matched.group(1)))


def euronext_instrument_type(name):
    upper = normalize_symbol(name)
    if any(token in upper for token in [" BSA", " BSAA", "WARRANT", "RIGHT", "DROIT"]):
        return "\u6743\u8bc1/\u6743\u5229"
    if "ETF" in upper or "FUND" in upper:
        return "ETF/\u57fa\u91d1\u7c7b\u6807\u7684"
    return "\u6b27\u6d32\u80a1\u7968"


def parse_euronext(text):
    reader = csv.reader(io.StringIO(text.lstrip("\ufeff")), delimiter=";")
    rows = list(reader)
    if not rows:
        return []
    header = [clean_name(item).lstrip("\ufeff") for item in rows[0]]
    parsed = []
    seen = set()
    for row in rows[1:]:
        if len(row) < len(header):
            continue
        record = dict(zip(header, row))
        symbol_raw = normalize_symbol(record.get("Symbol"))
        name = clean_name(record.get("Name"))
        market_name = clean_name(record.get("Market"))
        if not symbol_raw or not name or not market_name:
            continue
        suffix = EURONEXT_MARKET_SUFFIXES.get(market_name, "EU")
        symbol = f"{symbol_raw}.{suffix}"
        if symbol in seen:
            continue
        seen.add(symbol)
        isin = clean_name(record.get("ISIN"))
        aliases = [item for item in [symbol_raw, isin] if item]
        parsed.append({
            "symbol": symbol,
            "symbolRaw": symbol_raw,
            "aliases": aliases,
            "name": name,
            "market": "EURONEXT",
            "exchange": EURONEXT_MARKET_MICS.get(market_name, "EURONEXT"),
            "exchangeName": market_name,
            "country": "",
            "currency": clean_name(record.get("Currency")),
            "instrumentType": euronext_instrument_type(name),
            "isin": isin,
            "source": "Euronext product directory stocks-all-places CSV",
        })
    return parsed


def load_euronext_stocks():
    request = urllib.request.Request(
        euronext_download_url(),
        headers={
            "User-Agent": "Mozilla/5.0",
            "Referer": EURONEXT_EQUITIES_PAGE_URL,
        },
    )
    with urllib.request.urlopen(request, timeout=90) as response:
        return parse_euronext(response.read().decode("utf-8-sig", errors="replace"))


def xetra_instrument_type(record):
    segment = clean_name(record.get("Unit of Quotation")) or clean_name(record.get("Market Segment"))
    instrument_type = clean_name(record.get("Instrument Type"))
    lower = segment.lower()
    if "etf" in lower:
        return "ETF/\u57fa\u91d1\u7c7b\u6807\u7684"
    if "etc" in lower:
        return "ETC"
    if "etn" in lower:
        return "ETN"
    if "share" in lower:
        return "\u6b27\u6d32\u80a1\u7968"
    return segment or instrument_type or "\u6b27\u6d32\u8bc1\u5238"


def parse_xetra(text):
    reader = csv.reader(io.StringIO(text.lstrip("\ufeff")), delimiter=";")
    header = None
    parsed = []
    seen = set()
    for row in reader:
        if not row:
            continue
        cleaned = [clean_name(item).lstrip("\ufeff") for item in row]
        if not header and {"Instrument", "ISIN", "Mnemonic"}.issubset(set(cleaned)):
            header = cleaned
            continue
        if not header or len(cleaned) < len(header):
            continue
        record = dict(zip(header, cleaned))
        if record.get("Product Status") != "Active" or record.get("Instrument Status") != "Active":
            continue
        mnemonic = normalize_symbol(record.get("Mnemonic"))
        name = clean_name(record.get("Instrument"))
        if not mnemonic or not name:
            continue
        symbol = f"{mnemonic}.XETRA"
        if symbol in seen:
            continue
        seen.add(symbol)
        isin = clean_name(record.get("ISIN"))
        wkn = clean_name(record.get("WKN"))
        aliases = [item for item in [mnemonic, isin, wkn] if item]
        parsed.append({
            "symbol": symbol,
            "symbolRaw": mnemonic,
            "aliases": aliases,
            "name": name,
            "market": "XETRA",
            "exchange": clean_name(record.get("MIC Code")) or "XETR",
            "exchangeName": "Xetra",
            "country": clean_name(record.get("Country Of Issue")),
            "currency": clean_name(record.get("Currency")) or clean_name(record.get("Settlement Currency")),
            "instrumentType": xetra_instrument_type(record),
            "isin": isin,
            "source": "Deutsche B\u00f6rse Xetra all tradable instruments CSV",
        })
    return parsed


def load_xetra_instruments():
    return parse_xetra(download_text(xetra_download_url()))


def parse_six_equity_issuers(payload):
    parsed = []
    seen = set()
    for item in payload.get("itemList") or []:
        symbol_raw = normalize_symbol(item.get("valorSymbol"))
        name = clean_name(item.get("company"))
        isin = clean_name(item.get("isin"))
        if not symbol_raw or not name or not isin:
            continue
        symbol = f"{symbol_raw}.SW"
        if symbol in seen:
            continue
        seen.add(symbol)
        aliases = [item for item in [symbol_raw, isin, clean_name(item.get("valorNumber"))] if item]
        parsed.append({
            "symbol": symbol,
            "symbolRaw": symbol_raw,
            "aliases": aliases,
            "name": name,
            "market": "SIX",
            "exchange": clean_name(item.get("tradingPlatform")) or "XSWX",
            "exchangeName": "SIX Swiss Exchange",
            "country": clean_name(item.get("country")) or "CH",
            "currency": clean_name(item.get("tradingCurrency")) or "CHF",
            "instrumentType": clean_name(item.get("classOfShare")) or "欧洲股票",
            "isin": isin,
            "source": "SIX List of Equity Issuers API",
        })
    return parsed


def load_six_equity_issuers():
    payload = download_json_with_headers(
        SIX_EQUITY_ISSUERS_URL,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/json",
            "Referer": "https://www.six-group.com/en/market-data/shares/companies.html",
        },
    )
    return parse_six_equity_issuers(payload)


def bme_api_headers():
    return {
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json",
        "Origin": "https://www.bolsasymercados.es",
        "Referer": BME_LISTED_COMPANIES_PAGE_URL,
    }


def bme_api_url(path, params):
    return urllib.parse.urljoin(BME_MARKET_ENDPOINT, path) + "?" + urllib.parse.urlencode(params)


def load_bme_listed_companies(trading_system, mtf_segment=""):
    page = 0
    records = []
    while True:
        params = {
            "tradingSystem": trading_system,
            "page": page,
            "pageSize": 50,
        }
        if mtf_segment:
            params["mtfSegment"] = mtf_segment
        payload = download_json_with_headers(
            bme_api_url("v1/EQ/ListedCompanies", params),
            headers=bme_api_headers(),
            retries=4,
            retry_delay=1.0,
        )
        records.extend(payload.get("data") or [])
        if not payload.get("hasMoreResults"):
            break
        page += 1
    return records


def parse_bme_company_search(fallback, market_code, exchange_name):
    isin = clean_name(fallback.get("mainShareISIN") or fallback.get("isin"))
    name = clean_name(fallback.get("name") or fallback.get("shareName"))
    if not isin or not name:
        return None
    suffix = "MC" if market_code == "BME" else "BMEG"
    aliases = [
        item
        for item in [
            isin,
            clean_name(fallback.get("shareName")),
            clean_name(fallback.get("companyKey")),
        ]
        if item
    ]
    return {
        "symbol": f"{isin}.{suffix}",
        "symbolRaw": isin,
        "aliases": aliases,
        "name": name,
        "market": market_code,
        "exchange": "XMAD" if market_code == "BME" else "BMEG",
        "exchangeName": exchange_name,
        "country": "ES",
        "currency": "EUR",
        "instrumentType": "BME Growth 股票" if market_code == "BME_GROWTH" else "西班牙股票",
        "isin": isin,
        "source": f"{exchange_name} ListedCompanies API",
    }


def load_bme_companies():
    targets = [
        ("SIBE", "", "BME", "BME Main Market"),
        ("MTF", "BMEGrowth", "BME_GROWTH", "BME Growth"),
    ]
    parsed = []
    for trading_system, mtf_segment, market_code, exchange_name in targets:
        companies = load_bme_listed_companies(trading_system, mtf_segment)
        for company in companies:
            listing = parse_bme_company_search(company, market_code, exchange_name)
            if listing:
                parsed.append(listing)
    return parsed


def parse_lse_aim_instruments(text):
    reader = csv.DictReader(io.StringIO(text.replace("\r\n", "\n").replace("\r", "\n")))
    parsed = []
    seen = set()
    for record in reader:
        code = normalize_symbol(record.get("Code"))
        name = clean_name(record.get("Name"))
        if not code or not name:
            continue
        symbol = f"{code}.L"
        if symbol in seen:
            continue
        seen.add(symbol)
        parsed.append({
            "symbol": symbol,
            "symbolRaw": code,
            "aliases": [code],
            "name": name,
            "market": "LSE_AIM",
            "exchange": "XLON",
            "exchangeName": "London Stock Exchange AIM",
            "country": "GB",
            "currency": "",
            "instrumentType": "AIM 股票",
            "source": "London Stock Exchange AIM instruments CSV",
        })
    return parsed


def load_lse_aim_instruments():
    request = urllib.request.Request(
        LSE_AIM_INSTRUMENTS_URL,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "text/csv",
            "Origin": "https://www.londonstockexchange.com",
            "Referer": "https://www.londonstockexchange.com/reports?lang=en&tab=aim",
        },
    )
    with urllib.request.urlopen(request, timeout=90) as response:
        return parse_lse_aim_instruments(response.read().decode("utf-8-sig", errors="replace"))


NASDAQ_NORDIC_SUFFIXES = {
    "DK": ("CO", "XCSE", "Nasdaq Copenhagen"),
    "FI": ("HE", "XHEL", "Nasdaq Helsinki"),
    "IS": ("IC", "XICE", "Nasdaq Iceland"),
    "SE": ("ST", "XSTO", "Nasdaq Stockholm"),
}


def nasdaq_nordic_exchange(record):
    isin = clean_name(record.get("isin"))
    country_code = isin[:2].upper()
    return NASDAQ_NORDIC_SUFFIXES.get(country_code, ("NORDIC", "NASDAQ_NORDIC", "Nasdaq Nordic"))


def parse_nasdaq_nordic(payload, category):
    rows = ((payload.get("data") or {}).get("instrumentListing") or {}).get("rows") or []
    parsed = []
    seen = set()
    for record in rows:
        symbol_raw = normalize_symbol(record.get("symbol"))
        name = clean_name(record.get("fullName"))
        isin = clean_name(record.get("isin"))
        if not symbol_raw or not name or not isin:
            continue
        suffix, mic, exchange_name = nasdaq_nordic_exchange(record)
        symbol = f"{symbol_raw}.{suffix}"
        if symbol in seen:
            continue
        seen.add(symbol)
        aliases = [item for item in [symbol_raw, isin, clean_name(record.get("orderbookId"))] if item]
        parsed.append({
            "symbol": symbol,
            "symbolRaw": symbol_raw,
            "aliases": aliases,
            "name": name,
            "market": "NASDAQ_NORDIC",
            "exchange": mic,
            "exchangeName": f"{exchange_name} First North" if category == "FIRST_NORTH" else exchange_name,
            "country": isin[:2].upper(),
            "currency": clean_name(record.get("currency")),
            "instrumentType": "First North 股票" if category == "FIRST_NORTH" else "北欧股票",
            "isin": isin,
            "source": f"Nasdaq Nordic screener {category}",
        })
    return parsed


def load_nasdaq_nordic_shares():
    parsed = []
    headers = {"User-Agent": "Mozilla/5.0", "Accept": "application/json"}
    for category in ["MAIN_MARKET", "FIRST_NORTH"]:
        payload = download_json_with_headers(
            NASDAQ_NORDIC_SCREENER_URL.format(category=category),
            headers=headers,
            timeout=90,
        )
        parsed.extend(parse_nasdaq_nordic(payload, category))
    return parsed


def nordic_and_otc_seeds():
    return [
        {
            "symbol": "SIVE.ST",
            "symbolRaw": "SIVE",
            "name": "Sivers Semiconductors AB",
            "market": "STOCKHOLM",
            "exchange": "Nasdaq Stockholm",
            "exchangeName": "Nasdaq Stockholm",
            "country": "SE",
            "currency": "SEK",
            "instrumentType": "普通股",
            "source": "Nordic seed snapshot",
        },
        {
            "symbol": "SIVEF",
            "symbolRaw": "SIVEF",
            "name": "Sivers Semiconductors AB",
            "market": "OTC",
            "exchange": "OTC Markets",
            "exchangeName": "OTC Markets",
            "country": "US",
            "currency": "USD",
            "instrumentType": "OTC",
            "source": "OTC seed snapshot",
        },
    ]


def load_rwa_coverage():
    try:
        payload = json.loads(download_text(BINANCE_RWA_URL))
    except Exception as exc:
        print(f"RWA source failed: {exc}", file=sys.stderr)
        return {}
    coverage = defaultdict(lambda: {"symbols": set(), "chains": set(), "hasOndo": False, "hasXStock": False, "records": 0})
    for item in payload.get("data", []):
        ticker = normalize_symbol(item.get("ticker"))
        if not ticker:
            continue
        row = coverage[ticker]
        row["symbols"].add(item.get("symbol"))
        row["chains"].add(str(item.get("chainId")))
        row["hasOndo"] = row["hasOndo"] or int(item.get("type") or 0) == 1
        row["hasXStock"] = row["hasXStock"] or int(item.get("type") or 0) == 2
        row["records"] += 1
    return {
        ticker: {
            "symbols": sorted(v["symbols"]),
            "chains": sorted(v["chains"]),
            "hasOndo": v["hasOndo"],
            "hasXStock": v["hasXStock"],
            "records": v["records"],
        }
        for ticker, v in coverage.items()
    }


def load_yahoo_research(symbol, keywords):
    research = {
        "symbol": symbol,
        "quote": None,
        "news": [],
        "newsDiagnostics": {
            "source": "Yahoo Finance search + free-api-discovery probe",
            "accepted": 0,
            "rejected": 0,
            "note": "",
        },
    }
    try:
        payload = download_json(YAHOO_CHART_URL.format(symbol=urllib.parse.quote(symbol)))
        result = (payload.get("chart", {}).get("result") or [None])[0]
        meta = (result or {}).get("meta") or {}
        if meta:
            previous = meta.get("chartPreviousClose")
            price = meta.get("regularMarketPrice")
            change = None
            change_percent = None
            if isinstance(price, (int, float)) and isinstance(previous, (int, float)) and previous:
                change = price - previous
                change_percent = change / previous * 100
            market_time = meta.get("regularMarketTime")
            research["quote"] = {
                "source": "Yahoo Finance chart API",
                "symbol": meta.get("symbol") or symbol,
                "name": meta.get("longName") or meta.get("shortName") or symbol,
                "exchange": meta.get("fullExchangeName") or meta.get("exchangeName"),
                "currency": meta.get("currency"),
                "price": price,
                "previousClose": previous,
                "change": change,
                "changePercent": change_percent,
                "volume": meta.get("regularMarketVolume"),
                "marketTime": datetime.fromtimestamp(market_time, UTC).strftime("%Y-%m-%d %H:%M UTC") if market_time else "",
            }
    except Exception as exc:
        research["quoteError"] = str(exc)

    try:
        payload = download_json(YAHOO_SEARCH_URL.format(symbol=urllib.parse.quote(symbol)), user_agent="Mozilla/5.0")
        news = payload.get("news") or []
        lowered_keywords = [item.lower() for item in keywords]
        for item in news:
            title = clean_name(item.get("title"))
            related = [str(ticker).lower() for ticker in item.get("relatedTickers") or []]
            haystack = f"{title} {' '.join(related)}".lower()
            if any(keyword in haystack for keyword in lowered_keywords):
                research["news"].append({
                    "title": title,
                    "publisher": clean_name(item.get("publisher")),
                    "url": item.get("link"),
                    "publishedAt": datetime.fromtimestamp(item.get("providerPublishTime"), UTC).strftime("%Y-%m-%d %H:%M UTC") if item.get("providerPublishTime") else "",
                    "source": "Yahoo Finance search",
                })
            else:
                research["newsDiagnostics"]["rejected"] += 1
        research["newsDiagnostics"]["accepted"] = len(research["news"])
    except Exception as exc:
        research["newsDiagnostics"]["error"] = str(exc)
    return research


def main():
    temp = Path(gettempdir())
    nasdaq = parse_nasdaq_listed(download_text(NASDAQ_LISTED_URL))
    other = parse_other_listed(download_text(OTHER_LISTED_URL))
    warnings = []
    try:
        hkex = call_with_retries(load_hkex_equity_filter, attempts=3, retry_delay=2.0)
    except Exception as exc:
        warnings.append(f"HKEX equityfilter failed, fallback to ListOfSecurities.xlsx: {exc}")
        hkex_xlsx = temp / "traderoute_hkex_ListOfSecurities.xlsx"
        download_binary(HKEX_LIST_URL, hkex_xlsx)
        hkex = parse_hkex(hkex_xlsx)
    if len(hkex) < 1000:
        warnings.append(f"HKEX snapshot count is low: {len(hkex)}")
    try:
        euronext = load_euronext_stocks()
    except Exception as exc:
        warnings.append(f"Euronext stocks CSV failed: {exc}")
        euronext = []
    if euronext and len(euronext) < 1000:
        warnings.append(f"Euronext snapshot count is low: {len(euronext)}")
    try:
        xetra = load_xetra_instruments()
    except Exception as exc:
        warnings.append(f"Xetra all tradable instruments CSV failed: {exc}")
        xetra = []
    if xetra and len(xetra) < 4000:
        warnings.append(f"Xetra snapshot count is low: {len(xetra)}")
    try:
        six = load_six_equity_issuers()
    except Exception as exc:
        warnings.append(f"SIX equity issuers API failed: {exc}")
        six = []
    if six and len(six) < 200:
        warnings.append(f"SIX equity issuer count is low: {len(six)}")
    try:
        bme = load_bme_companies()
    except Exception as exc:
        warnings.append(f"BME ListedCompanies API failed: {exc}")
        bme = []
    if bme and len(bme) < 200:
        warnings.append(f"BME SIBE/BME Growth count is low: {len(bme)}")
    try:
        lse_aim = load_lse_aim_instruments()
    except Exception as exc:
        warnings.append(f"LSE AIM instruments CSV failed: {exc}")
        lse_aim = []
    if lse_aim and len(lse_aim) < 500:
        warnings.append(f"LSE AIM instrument count is low: {len(lse_aim)}")
    warnings.append("LSE Main Market full issuer list remains a supplement candidate; no stable official full-list endpoint is wired.")
    try:
        nasdaq_nordic = load_nasdaq_nordic_shares()
    except Exception as exc:
        warnings.append(f"Nasdaq Nordic screener failed: {exc}")
        nasdaq_nordic = []
    if nasdaq_nordic and len(nasdaq_nordic) < 500:
        warnings.append(f"Nasdaq Nordic screener count is low: {len(nasdaq_nordic)}")
    seeds = nordic_and_otc_seeds()
    rwa_coverage = load_rwa_coverage()

    stocks_by_symbol = {}
    source_counter = defaultdict(int)
    for listing in [*nasdaq, *other, *hkex, *euronext, *xetra, *six, *bme, *lse_aim, *nasdaq_nordic, *seeds]:
        symbol = listing["symbol"]
        stock = row_to_stock(listing, listing["source"])
        if symbol in rwa_coverage:
            stock["profile"]["rwaSymbols"] = rwa_coverage[symbol]["symbols"]
        stocks_by_symbol[symbol] = stock
        source_counter[listing["source"]] += 1

    for ticker, coverage in rwa_coverage.items():
        if ticker not in stocks_by_symbol:
            listing = {
                "symbol": ticker,
                "symbolRaw": ticker,
                "name": PROFILE_OVERRIDES.get(ticker, {}).get("name") or f"{ticker} tokenized stock",
                "market": "US",
                "exchange": "RWA Snapshot",
                "exchangeName": "RWA Snapshot",
                "country": "US",
                "currency": "USD",
                "instrumentType": "RWA 快照标的",
                "source": "Binance Web3 RWA public API",
            }
            stock = row_to_stock(listing, listing["source"])
            stock["profile"]["rwaSymbols"] = coverage["symbols"]
            stocks_by_symbol[ticker] = stock

    for stock in list(stocks_by_symbol.values()):
        for related_symbol in stock["profile"].get("relatedListings", []):
            related = stocks_by_symbol.get(related_symbol)
            if not related:
                continue
            for listing in related["listings"]:
                if listing["symbol"] not in {item["symbol"] for item in stock["listings"]}:
                    stock["listings"].append(listing)
            if related_symbol not in stock["symbols"]:
                stock["symbols"].append(related_symbol)
            existing_routes = {route["platformId"] + "|" + route["listingSymbol"] for route in stock["profile"].get("verifiedBrokerRoutes", [])}
            for route in related["profile"].get("verifiedBrokerRoutes", []):
                route_key = route["platformId"] + "|" + route["listingSymbol"]
                if route_key not in existing_routes:
                    stock["profile"].setdefault("verifiedBrokerRoutes", []).append(route)
                    existing_routes.add(route_key)

    research_cache = {}
    for stock in stocks_by_symbol.values():
        override = PROFILE_OVERRIDES.get(stock["profile"]["code"], {})
        research_symbol = override.get("researchSymbol")
        if not research_symbol:
            continue
        if research_symbol not in research_cache:
            research_cache[research_symbol] = load_yahoo_research(research_symbol, override.get("researchKeywords", []))
        stock["profile"]["research"] = research_cache[research_symbol]
        if override.get("newsProbeNote"):
            stock["profile"]["research"]["newsDiagnostics"]["note"] = override["newsProbeNote"]

    stocks = sorted(stocks_by_symbol.values(), key=lambda item: (item["profile"]["primaryMarket"], item["profile"]["code"]))
    output = {
        "meta": {
            "generatedAt": datetime.now().strftime("%Y-%m-%d"),
            "schemaVersion": "2.4.1",
            "stockCount": len(stocks),
            "listingCount": sum(len(item["listings"]) for item in stocks),
            "rwaTickerCount": len(rwa_coverage),
            "sources": dict(sorted(source_counter.items())),
            "warnings": warnings,
            "strategy": "security_master_with_official_exchange_lists_market_coverage_route_candidates_and_exposure_adapters",
        },
        "stocks": stocks,
        "platforms": {
            "brokers": BROKER_PLATFORMS,
            "crypto": CRYPTO_PLATFORMS,
        },
        "rwaCoverage": rwa_coverage,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(output, ensure_ascii=False, separators=(",", ":"))
    OUTPUT.write_text("window.MARKET_DATA = " + payload + ";\n", encoding="utf-8")
    print(f"Generated {len(stocks)} stocks, {output['meta']['listingCount']} listings.")


if __name__ == "__main__":
    main()
